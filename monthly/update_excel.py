# -*- coding: utf-8 -*-
"""
網站檢核表 Excel 自動更新

流程: 深掃產出 compliance.json 後, 跑本程式產生新月份的檢核表。

用法:
  python update_excel.py                     自動尋找上月檢核表, 產生本月新檔
  python update_excel.py --src <路徑.xlsx>   指定來源檔
  python update_excel.py --rename-tabs       一次性：把舊分頁名（工作表名稱）改成網站名稱
  python update_excel.py --dry-run           只顯示會做什麼, 不寫檔

動作:
  1. 複製上月檢核表 → 「{本月}資訊局網站檢核表（數據範圍{上月}01~{上月}月底）.xlsx」
  2. 每張網站工作表(以「網站名稱」為鍵):
     - 填表日期(E2) 更新為今天
     - 流量數(E4) 保留上月數字並標黃 → 待人工填入新數據
     - (二)檢索/(三)HTTPS/(四)RWD: 與自動檢測結果比對
       一致 → 不動 | 不一致 → 標黃並列入待確認清單 (不自動改寫判定)
  3. 產出 reports/待人工確認_{本月}.md
  (流量總表為公式自動連動, 不需處理)
"""
import calendar
import datetime
import glob
import json
import os
import re
import sys

import openpyxl
from openpyxl.styles import PatternFill

import config

FULL_OVERNIGHT_GLOB = os.path.join(config.PRIVATE_DIR, "reports", "full_overnight_*")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
YELLOW = PatternFill("solid", start_color="FFFF00")

ROW_KEYS = {  # 工作表A欄關鍵字 → 檢測項目
    "search": "檢索功能是否完善",
    "https": "(三)網站使用HTTPS",
    "rwd": "(四)網站設計應符合響應式",
}


def roc_today():
    t = datetime.date.today()
    return t.year - 1911, t.month, t.day


def months():
    """回傳 (本月標籤, 上月標籤, 上月資料範圍起, 迄) 如 ('11506','11505','1150501','1150531')"""
    y, m, _ = roc_today()
    py, pm = (y, m - 1) if m > 1 else (y - 1, 12)
    last = calendar.monthrange(py + 1911, pm)[1]
    return (f"{y}{m:02d}", f"{py}{pm:02d}",
            f"{py}{pm:02d}01", f"{py}{pm:02d}{last:02d}")


def find_src():
    """尋找最新的檢核表檔案當來源 (排除備份/暫存檔)。
    搜尋: 檢核表/<年>/ 年度資料夾、本資料夾、Downloads"""
    cands = glob.glob(os.path.join(config.CHECKLIST_DIR, "*", "*資訊局網站檢核表*.xlsx"))
    for d in (os.path.join(os.environ.get("USERPROFILE", ""), "Downloads"), BASE_DIR):
        cands += glob.glob(os.path.join(d, "*資訊局網站檢核表*.xlsx"))
    cands = [c for c in cands if "備份" not in c and "~$" not in c]
    if not cands:
        sys.exit("找不到來源檢核表, 請用 --src 指定")
    return max(cands, key=os.path.getmtime)


def parse_mark(v):
    """儲存格判定值 → True(○)/False(✕)/None(無法判讀)"""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    head = s[0]
    if head in "○Ｏo〇0◯⭕✓✔√" or head.upper() == "O" or head in "是有":
        return True
    if head in "✕XｘＸx╳╳❌✗╳" or head in "否無":
        return False
    return None


def site_detection(res):
    """彙整一個網站(可能多URL)的檢測結果 → dict(search/https/rwd: bool|None)"""
    out = {"search": None, "https": None, "rwd": None, "broken": [], "alive": True}
    for url, r in res.get("urls", {}).items():
        if not r.get("alive"):
            out["alive"] = False
        cert_ok = r.get("https", {}).get("cert", {}).get("valid")
        out["https"] = (out["https"] is not False) and bool(cert_ok)
        rwd = r.get("rwd")
        if rwd is not None:
            ok = rwd.get("viewport") and rwd.get("responsive_css")
            out["rwd"] = (out["rwd"] is not False) and bool(ok)
        if r.get("search") is not None:
            out["search"] = bool(out["search"]) or bool(r["search"])
        out["broken"] += [f"{u} → {why}" for u, why in r.get("links_broken", [])]
    return out


def find_compliance_json():
    """找最新深掃目錄的 compliance.json。"""
    dirs = sorted(glob.glob(FULL_OVERNIGHT_GLOB), reverse=True)
    for d in dirs:
        p = os.path.join(d, "compliance.json")
        if os.path.exists(p):
            return p
    return None


def _sanitize_tab_name(name, max_len=31):
    """Excel 分頁名限 31 字、不可含 \\/:*?[]。"""
    name = re.sub(r'[\\/:*?\[\]]', '_', name).strip()
    return name[:max_len] if name else "_"


def rename_tabs(wb, dry=False):
    """一次性：把舊分頁名（工作表名稱）改成對應的網站名稱。
    對照優先讀 sites.json 的 sheet 欄；該欄已隨主設定表退役(新 sync 後為空)，
    故備援讀重構時固化的 private/tab_rename_map.json。"""
    mapping = {}
    if os.path.exists(config.SITES_JSON):
        with open(config.SITES_JSON, encoding="utf-8") as f:
            mapping = {s["sheet"]: s["name"] for s in json.load(f)["sites"]
                       if s.get("sheet") and s.get("name")}
    if not mapping:
        snap = os.path.join(config.PRIVATE_DIR, "tab_rename_map.json")
        if os.path.exists(snap):
            mapping = json.load(open(snap, encoding="utf-8"))
    if not mapping:
        print("  (略過分頁改名：sites.json 無 sheet 欄且無 tab_rename_map.json)")
        return []
    renamed = []
    for old_name in list(wb.sheetnames):
        if old_name in mapping:
            new_name = _sanitize_tab_name(mapping[old_name])
            if new_name != old_name and new_name not in wb.sheetnames:
                if not dry:
                    wb[old_name].title = new_name
                renamed.append((old_name, new_name))
                print(f"  分頁改名: {old_name} → {new_name}")
    return renamed


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    args = sys.argv[1:]
    dry = "--dry-run" in args
    src = args[args.index("--src") + 1] if "--src" in args else find_src()

    this_m, data_m, d_from, d_to = months()

    # 讀合規結果（來自深掃 compliance.json，以網站名稱為鍵）
    comp_path = find_compliance_json()
    if not comp_path:
        sys.exit("找不到 compliance.json\n請先執行深掃: python -m engine.full_overnight")
    print(f"合規數據: {comp_path}")
    with open(comp_path, encoding="utf-8") as f:
        results = json.load(f)   # {網站名稱: {urls:{...}, ai:[...]}}
    with open(config.SITES_JSON, encoding="utf-8") as f:
        site_cfg = {s["name"]: s for s in json.load(f)["sites"]}

    out_name = f"{this_m}資訊局網站檢核表（數據範圍{d_from}~{d_to}）.xlsx"
    # 比照 Drive 結構：本機輸出放「檢核表/<民國年>年/」資料夾
    year_dir = os.path.join(config.CHECKLIST_DIR, f"{this_m[:3]}年")
    os.makedirs(year_dir, exist_ok=True)
    out_path = os.path.join(year_dir, out_name)
    print(f"來源: {src}\n輸出: {out_path}\n檢測數據: {result_path}\n")

    wb = openpyxl.load_workbook(src)
    # 一次性分頁改名（--rename-tabs 或首次偵測到舊名時自動執行）
    if "--rename-tabs" in args:
        rename_tabs(wb, dry=dry)

    y, m, d = roc_today()
    todo = [f"# {this_m} 檢核表 待人工確認清單", "",
            f"產生時間：{datetime.datetime.now():%Y-%m-%d %H:%M}",
            f"已自動更新：各表填表日期 → {y} 年 {m} 月 {d} 日", ""]

    for site_name, res in results.items():
        if site_name not in wb.sheetnames:
            todo.append(f"## {site_name}\n- ⚠ 工作表不存在於來源檔，請確認")
            continue
        ws = wb[site_name]
        items = []

        # 1. 填表日期
        e2 = str(ws["E2"].value or "")
        new_e2 = re.sub(r"\d+\s*年\s*\d+\s*月\s*\d*\s*日?",
                        f"{y} 年 {m} 月 {d}日", e2) if e2 else f"填表日期：{y} 年 {m} 月 {d}日"
        if not dry:
            ws["E2"] = new_e2

        # 1b. 表頭網站名稱
        cfg = site_cfg.get(site_name, {})
        b2_val = str(cfg.get("name", "") or "").strip()
        if b2_val:
            old = str(ws["B2"].value or "")
            newv = "網站名稱：" + b2_val
            if old != newv and not dry:
                ws["B2"] = newv

        # 2. 流量數: 有設定 GA 資源者自動撈取, 其餘標黃人工填
        ga_pid = cfg.get("ga_property")
        traffic_done = False
        if ga_pid:
            import ga_traffic
            ga_metric = cfg.get("ga_metric", "screenPageViews")
            start, end = ga_traffic.roc_month_range(data_m)
            n, err = ga_traffic.fetch_pageviews(ga_pid, start, end, ga_metric)
            if err is None:
                old_e4 = str(ws["E4"].value or "")
                note = "【註】" + old_e4.split("【註】", 1)[1] if "【註】" in old_e4 else ""
                if not dry:
                    ws["E4"] = f"網站流量數:{n:,}\n{note}"
                    ws["E4"].fill = PatternFill()  # 自動填入成功, 清除待填標記
                items.append(f"- ✅ E4 流量數：已自動填入 GA 數據 {n:,}（{start}~{end}）")
                traffic_done = True
            else:
                items.append(f"- ⚠ E4 流量數：GA 撈取失敗（{err[:80]}），請人工填入")
        if not traffic_done:
            if not dry:
                ws["E4"].fill = YELLOW
            if not ga_pid:
                items.append(f"- 🟡 E4 流量數：請填入 {data_m} 月份數據（目前仍是上月數字）")

        # 3. 比對 (二)(三)(四)
        det = site_detection(res)
        if not det["alive"]:
            items.append("- ❌ 網站連線異常，請人工確認後再填表")
        label = {"search": "(二)檢索功能", "https": "(三)HTTPS", "rwd": "(四)RWD"}
        for key, kw in ROW_KEYS.items():
            if det[key] is None:
                continue
            for row in ws.iter_rows(min_col=1, max_col=1):
                a = row[0].value
                if a and kw in str(a):
                    cell = ws.cell(row=row[0].row, column=4)  # D欄
                    old = parse_mark(cell.value)
                    if old is not None and old != det[key]:
                        if not dry:
                            cell.fill = YELLOW
                        items.append(
                            f"- 🟡 D{row[0].row} {label[key]}：表上填「{str(cell.value)[:20]}」"
                            f"，但自動檢測為「{'○' if det[key] else '✕'}」，請確認後修正")
                    break

        # 4. 失效連結
        for b in det["broken"]:
            items.append(f"- ❌ 失效連結：{b}（影響(一)超連結有效性）")

        # 4b. 站內深度檢測結果（來自 compliance deep_check）
        for _url, _r in res.get("urls", {}).items():
            deep = _r.get("deep") or {}
            if deep.get("broken_internal"):
                bi = deep["broken_internal"]
                mailto_n = sum(1 for u, *_ in bi
                               if "mailto:" in u.lower() or "tel:" in u.lower())
                real_n = len(bi) - mailto_n
                seg = []
                if real_n:
                    seg.append(f"缺頁 {real_n} 筆")
                if mailto_n:
                    seg.append(f"網站方 email/電話連結誤寫成相對路徑 {mailto_n} 筆")
                items.append(f"- ❌ 站內失效頁面 {len(bi)} 筆（{'、'.join(seg)}；"
                             "見檢測報告，影響(一)超連結有效性）")
            if deep.get("office_no_universal"):
                items.append(f"- 📎 下載文件未提供PDF/ODF通用格式 {len(deep['office_no_universal'])} 筆"
                             "（見檢測報告，影響(一)下載文件通用格式）")

        # 5. AI 判讀摘要
        for ai in res.get("ai", []):
            ans = ai["answer"].replace("\n", " ")[:120]
            items.append(f"- 🤖 AI判讀 {ai['url']}：{ans}")

        todo += [f"## {site_name}"] + items + [""]

    if not dry:
        wb.save(out_path)
        todo_path = os.path.join(config.REPORTS_DIR, f"待人工確認_{this_m}.md")
        with open(todo_path, "w", encoding="utf-8") as f:
            f.write("\n".join(todo))
        print(f"已產生新檢核表: {out_path}")
        print(f"待確認清單:     {todo_path}")
    else:
        print("\n".join(todo))


if __name__ == "__main__":
    main()
