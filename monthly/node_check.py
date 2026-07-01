# -*- coding: utf-8 -*-
"""
檢核表(一) 逐節點 AI 內容判讀

兩步驟:
  python node_check.py map [檢核表.xlsx]   從檢核表抽出各站(一)的節點(填○者),
                                           爬各站選單自動對應節點→網址,
                                           產生 nodes_map.json(可手動修正補漏)
  python node_check.py check [--month 11506]
                                           逐節點抓頁面請地端AI判讀內容是否過期/錯誤,
                                           產生 reports/節點判讀_<月份>.md

對應不到網址的節點會列在 nodes_map.json 的 url=null, 請人工補上後重跑 check。
"""
import glob
import json
import os
import re
import sys
import urllib.parse
from difflib import SequenceMatcher

import openpyxl

import config
import webcheck_ai

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MAP_PATH = config.NODES_MAP
CRAWL_PAGES = 30  # 建對應表時每站爬的頁數(首頁+主要內頁的選單)


def parse_mark(v):
    s = str(v or "").strip()
    if not s:
        return None
    if s[0] in "○Ｏo〇" or s[0].upper() == "O":
        return True
    return False


def extract_nodes(xlsx_path, sheet_names):
    """從檢核表抽出各站(一)中檢核為○的節點清單"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = {}
    for sheet in sheet_names:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        nodes = []
        level1 = ""
        for row in ws.iter_rows(min_row=10, max_row=120):
            a = str(row[0].value or "").strip()
            if a.startswith("【註】") or a.startswith("(二)"):
                break
            b = str(row[1].value or "").strip()
            if a and not a.startswith("【"):
                level1 = a
            name = b or a
            if not name:
                continue
            if parse_mark(row[3].value):  # D欄=內容正確有效, 填○才需逐月判讀
                nodes.append({"level1": re.sub(r"\s+", "", level1),
                              "name": re.sub(r"\s+", "", name)})
        out[sheet] = nodes
    return out


def name_candidates(name):
    """節點名稱 → 比對候選詞: 去括號註記、括號內實際名稱單獨成詞"""
    cands = []
    base = re.sub(r"[（(].*?[）)]", "", name).strip()
    if base:
        cands.append(base)
    for m in re.findall(r"[（(](.+?)[）)]", name):
        m = m.strip()
        if m and m not in ("-",):
            cands.append(m)
    return cands or [name]


def crawl_anchors(start_url, max_pages=CRAWL_PAGES):
    """爬站內頁面收集 連結文字→網址 對照(取最常見的對應)"""
    import requests
    from bs4 import BeautifulSoup
    requests.packages.urllib3.disable_warnings()
    sess = requests.Session()
    sess.headers.update({"User-Agent": webcheck_ai.UA, "Accept": "text/html,*/*;q=0.8"})
    host = urllib.parse.urlsplit(start_url).hostname
    seen = {start_url}
    frontier = [start_url]
    anchors = {}  # text -> url (首見優先)
    done = 0
    while frontier and done < max_pages:
        page = frontier.pop(0)
        try:
            r = sess.get(page, timeout=15, verify=True)  # 預設驗證憑證
        except requests.exceptions.SSLError:
            try:  # 憑證鏈不完整 → 退回不驗證
                r = sess.get(page, timeout=15, verify=False)
            except Exception:
                continue
        except Exception:
            continue
        done += 1
        if "html" not in r.headers.get("Content-Type", ""):
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.find_all("a", href=True):
            text = re.sub(r"\s+", "", a.get_text(strip=True))[:40]
            absu = urllib.parse.urldefrag(urllib.parse.urljoin(r.url, a["href"].strip()))[0]
            if not text or not absu.lower().startswith(("http://", "https://")):
                continue
            anchors.setdefault(text, absu)
            h = urllib.parse.urlsplit(absu).hostname
            if h == host and absu not in seen and len(seen) < 300:
                seen.add(absu)
                frontier.append(absu)
    return anchors


def match_node(name, anchors):
    """節點名稱對應網址: 完全相符 > 互相包含 > 相似度>0.7"""
    cands = name_candidates(name)
    for c in cands:
        if c in anchors:
            return anchors[c]
    for c in cands:
        for text, url in anchors.items():
            if c and (c in text or text in c) and len(text) >= 2:
                return url
    best, best_r = None, 0.7
    for c in cands:
        for text, url in anchors.items():
            r = SequenceMatcher(None, c, text).ratio()
            if r > best_r:
                best, best_r = url, r
    return best


def build_map(xlsx_path):
    with open(config.SITES_JSON, encoding="utf-8") as f:
        sites = json.load(f)["sites"]
    sheet_urls = {s["sheet"]: s["urls"][0] for s in sites}
    nodes_by_sheet = extract_nodes(xlsx_path, list(sheet_urls))
    result = {}
    for sheet, nodes in nodes_by_sheet.items():
        if not nodes:
            continue
        url = sheet_urls[sheet]
        print(f"{sheet}: {len(nodes)} 個○節點, 爬取選單對應中...")
        anchors = crawl_anchors(url)
        mapped = 0
        for n in nodes:
            n["url"] = match_node(n["name"], anchors)
            mapped += bool(n["url"])
        print(f"  自動對應 {mapped}/{len(nodes)}")
        result[sheet] = nodes
    with open(MAP_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"\n對照表已存 {MAP_PATH}, 未對應(url=null)請人工補上")


AI_QUESTION = ("此頁面是機關網站的「{node}」單元。請依網站檢核標準判讀並簡short回答: "
               "1.內容是否有明顯錯誤、亂碼或過期資訊(若頁面有日期請註明最新一筆日期) "
               "2.標題與內容是否一致 "
               "3.綜合建議: 填○(正常)或✕(異常, 說明原因)")


def run_check(month, sheet_filter=None):
    if not os.path.exists(MAP_PATH):
        sys.exit("找不到 nodes_map.json, 請先執行: python node_check.py map <檢核表.xlsx>")
    with open(MAP_PATH, encoding="utf-8") as f:
        node_map = json.load(f)
    if sheet_filter:
        node_map = {k: v for k, v in node_map.items() if sheet_filter in k}
    # 各站抓取方式(來自主設定表，經 sync 寫入 sites.json)
    methods = {}
    sp = config.SITES_JSON
    if os.path.exists(sp):
        methods = {s["sheet"]: s.get("method", "ai") for s in json.load(open(sp, encoding="utf-8"))["sites"]}
    escalate = []  # 建議升級成 playwright 的站
    lines = [f"# {month} 檢核表(一) 逐節點 AI 內容判讀", ""]
    total = sum(len(v) for v in node_map.values())
    done = 0
    for sheet, nodes in node_map.items():
        method = methods.get(sheet, "ai")
        lines.append(f"## {sheet}（抓取方式：{method}）")
        if method == "manual":
            lines.append("- ⏭ 此站列為人工檢視（自動化讀不到內容，如3D地圖），逐節點請人工確認")
            lines.append("")
            done += len(nodes)
            continue
        for n in nodes:
            done += 1
            label = f"{n['level1']}>{n['name']}" if n.get("level1") else n["name"]
            print(f"[{done}/{total}] {sheet} {label}", flush=True)
            if not n.get("url"):
                lines.append(f"- ⚠ {label}：無對應網址，請人工檢視（可在 nodes_map.json 補上）")
                continue
            try:
                text, used, note = webcheck_ai.get_content(n["url"], method)
                ans = webcheck_ai.ask_ai(AI_QUESTION.format(node=n["name"]), text, n["url"])
                ans = ans.strip().replace("\n", " ¶ ")
                if note:
                    ans += f" ¶ [{note}]"
                # ai/code 抓到空內容 → 建議升級 playwright
                if method in ("ai", "code") and webcheck_ai.looks_empty(ans) and sheet not in escalate:
                    escalate.append(sheet)
            except Exception as e:
                ans = f"(判讀失敗: {type(e).__name__}: {e})"
            mark = "✕" if "✕" in ans else ("○" if "○" in ans else "?")
            lines.append(f"- {mark} **{label}** `{n['url']}`\n  - {ans}")
        lines.append("")
    if escalate:
        lines.append("## ⚙ 建議調整抓取方式")
        for s in escalate:
            lines.append(f"- {s}：AI 在靜態內容抓到「空/查無」，建議在主設定表把「內容抓取方式」改為 **playwright**")
    out = os.path.join(config.REPORTS_DIR, f"節點判讀_{month}.md")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"\n完成: {out}")


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        return
    if args[0] == "map":
        if len(args) > 1:
            xlsx = args[1]
        else:
            cands = sorted(glob.glob(os.path.join(config.CHECKLIST_DIR, "*", "*資訊局網站檢核表*.xlsx")))
            if not cands:
                sys.exit("請指定檢核表路徑: python node_check.py map <xlsx>")
            xlsx = cands[-1]
        print("使用檢核表:", xlsx)
        build_map(xlsx)
    elif args[0] == "check":
        if "--month" in args:
            month = args[args.index("--month") + 1]
        else:
            import datetime
            t = datetime.date.today()
            month = f"{t.year - 1911}{t.month:02d}"
        sheet_filter = args[args.index("--sheet") + 1] if "--sheet" in args else None
        run_check(month, sheet_filter)


if __name__ == "__main__":
    main()
